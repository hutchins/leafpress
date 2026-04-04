"""Tests for XLSX to Markdown import converter."""

from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path

import pytest
from openpyxl import Workbook

from leafpress.exceptions import XlsxImportError
from leafpress.importer.converter_xlsx import import_xlsx

# --- helpers ---


def _make_xlsx(tmp_path: Path, sheets: dict[str, list[list]]) -> Path:
    """Create an XLSX file from sheet definitions.

    Args:
        sheets: Mapping of sheet name → list of rows (each row a list of cell values).
    """
    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)

    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)

    path = tmp_path / "test.xlsx"
    wb.save(path)
    return path


# --- import_xlsx tests ---


def test_import_xlsx_basic(tmp_path: Path) -> None:
    """Single sheet with header + data rows produces a Markdown table."""
    xlsx_path = _make_xlsx(
        tmp_path,
        {
            "Data": [
                ["Name", "Value"],
                ["Alpha", 100],
                ["Beta", 200],
            ],
        },
    )
    result = import_xlsx(xlsx_path)
    md = result.markdown_path.read_text()
    assert "## Data" in md
    assert "| Name" in md
    assert "| Alpha" in md
    assert "| Beta" in md
    assert "---" in md  # separator row
    assert "100" in md
    assert "200" in md


def test_import_xlsx_multiple_sheets(tmp_path: Path) -> None:
    """Each sheet gets its own ## heading and table."""
    xlsx_path = _make_xlsx(
        tmp_path,
        {
            "Users": [
                ["Name", "Email"],
                ["Alice", "alice@example.com"],
            ],
            "Products": [
                ["SKU", "Price"],
                ["A001", 9.99],
            ],
        },
    )
    result = import_xlsx(xlsx_path)
    md = result.markdown_path.read_text()
    assert "## Users" in md
    assert "## Products" in md
    assert "Alice" in md
    assert "A001" in md


def test_import_xlsx_empty_sheet_skipped(tmp_path: Path) -> None:
    """Empty worksheets are not included in output."""
    xlsx_path = _make_xlsx(
        tmp_path,
        {
            "HasData": [["Col1"], ["Val1"]],
            "Empty": [],
        },
    )
    result = import_xlsx(xlsx_path)
    md = result.markdown_path.read_text()
    assert "## HasData" in md
    assert "Empty" not in md


def test_import_xlsx_empty_cells(tmp_path: Path) -> None:
    """Blank cells render as empty table cells."""
    xlsx_path = _make_xlsx(
        tmp_path,
        {
            "Sparse": [
                ["A", "B", "C"],
                [1, None, 3],
                [None, 2, None],
            ],
        },
    )
    result = import_xlsx(xlsx_path)
    md = result.markdown_path.read_text()
    assert "## Sparse" in md
    # The table should still have proper pipe structure
    lines = [line for line in md.split("\n") if line.startswith("|")]
    assert len(lines) == 4  # header + separator + 2 data rows


def test_import_xlsx_numeric_and_date_values(tmp_path: Path) -> None:
    """Numbers and dates render as strings."""
    xlsx_path = _make_xlsx(
        tmp_path,
        {
            "Types": [
                ["Type", "Value"],
                ["Integer", 42],
                ["Float", 3.14],
                ["Date", date(2025, 6, 15)],
                ["DateTime", datetime(2025, 6, 15, 10, 30, 0)],
            ],
        },
    )
    result = import_xlsx(xlsx_path)
    md = result.markdown_path.read_text()
    assert "42" in md
    assert "3.14" in md
    assert "2025-06-15" in md
    assert "10:30:00" in md


def test_import_xlsx_pipe_in_cell_escaped(tmp_path: Path) -> None:
    """Pipe characters in cell values are escaped."""
    xlsx_path = _make_xlsx(
        tmp_path,
        {
            "Pipes": [
                ["Command", "Description"],
                ["a | b", "pipe example"],
            ],
        },
    )
    result = import_xlsx(xlsx_path)
    md = result.markdown_path.read_text()
    assert "a \\| b" in md


def test_import_xlsx_output_path_default(tmp_path: Path) -> None:
    """Default output uses same stem as input."""
    xlsx_path = _make_xlsx(tmp_path, {"Sheet1": [["A"], [1]]})
    result = import_xlsx(xlsx_path)
    assert result.markdown_path == tmp_path / "test.md"


def test_import_xlsx_output_path_directory(tmp_path: Path) -> None:
    """Directory output creates <stem>.md inside it."""
    xlsx_path = _make_xlsx(tmp_path, {"Sheet1": [["A"], [1]]})
    out_dir = tmp_path / "out"
    out_dir.mkdir()
    result = import_xlsx(xlsx_path, output_path=out_dir)
    assert result.markdown_path == out_dir / "test.md"


def test_import_xlsx_output_path_explicit(tmp_path: Path) -> None:
    """Explicit output path is respected."""
    xlsx_path = _make_xlsx(tmp_path, {"Sheet1": [["A"], [1]]})
    out = tmp_path / "custom" / "output.md"
    result = import_xlsx(xlsx_path, output_path=out)
    assert result.markdown_path == out
    assert out.exists()


def test_import_xlsx_missing_file(tmp_path: Path) -> None:
    """Nonexistent file raises XlsxImportError."""
    with pytest.raises(XlsxImportError, match="File not found"):
        import_xlsx(tmp_path / "nonexistent.xlsx")


def test_import_xlsx_wrong_extension(tmp_path: Path) -> None:
    """Wrong extension raises XlsxImportError."""
    txt_file = tmp_path / "data.csv"
    txt_file.write_text("a,b\n1,2\n")
    with pytest.raises(XlsxImportError, match=r"Not a \.xlsx file"):
        import_xlsx(txt_file)


def test_import_xlsx_no_images(tmp_path: Path) -> None:
    """XLSX import never produces images."""
    xlsx_path = _make_xlsx(tmp_path, {"Sheet1": [["A"], [1]]})
    result = import_xlsx(xlsx_path)
    assert result.images == []


def test_import_xlsx_integer_floats(tmp_path: Path) -> None:
    """Whole-number floats display without decimal point."""
    xlsx_path = _make_xlsx(
        tmp_path,
        {"Numbers": [["Val"], [1.0], [42.0], [3.5]]},
    )
    result = import_xlsx(xlsx_path)
    md = result.markdown_path.read_text()
    # 1.0 and 42.0 should display as "1" and "42"
    assert "| 1 " in md or "| 1  " in md
    assert "| 42" in md
    assert "3.5" in md


# --- CLI integration ---


def test_cli_xlsx_integration(tmp_path: Path) -> None:
    """End-to-end CLI import of an XLSX file."""
    from typer.testing import CliRunner

    from leafpress.cli import cli

    xlsx_path = _make_xlsx(
        tmp_path,
        {"Report": [["Metric", "Value"], ["Revenue", 1000]]},
    )
    runner = CliRunner()
    result = runner.invoke(cli, ["import", str(xlsx_path), "-o", str(tmp_path / "out.md")])
    assert result.exit_code == 0
    assert "Done!" in result.output
    assert (tmp_path / "out.md").exists()
    md = (tmp_path / "out.md").read_text()
    assert "## Report" in md
    assert "Revenue" in md


# ---------------------------------------------------------------------------
# Comprehensive XLSX fixture — realistic multi-sheet workbook
# ---------------------------------------------------------------------------


def _make_comprehensive_xlsx(tmp_path: Path) -> Path:
    """Create a multi-sheet XLSX with diverse data types and edge cases."""
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: "Summary" — mixed data types
    ws1 = wb.create_sheet(title="Summary")
    ws1.append(["Metric", "Value", "Date", "Time"])
    ws1.append(["Revenue", 2400000, date(2024, 12, 31), time(9, 30, 0)])
    ws1.append(["Expenses", 1800000.50, datetime(2024, 12, 31, 17, 0, 0), time(17, 0, 0)])
    ws1.append(["Headcount", 150.0, date(2025, 1, 15), None])

    # Sheet 2: "Details" — pipe chars, empty cells, whole-number floats
    ws2 = wb.create_sheet(title="Details")
    ws2.append(["Command", "Description", "Status"])
    ws2.append(["a | b", "pipe example", "active"])
    ws2.append([None, "missing command", "pending"])
    ws2.append(["ls -la", None, "done"])
    ws2.append(["echo hi", "simple", 42.0])

    # Sheet 3: "Empty" — no data at all
    wb.create_sheet(title="Empty")

    # Sheet 4: "Trailing" — data rows followed by empty rows
    ws4 = wb.create_sheet(title="Trailing")
    ws4.append(["Key", "Value"])
    ws4.append(["alpha", 1])
    ws4.append(["beta", 2])
    ws4.append([None, None])
    ws4.append([None, None])
    ws4.append([None, None])

    path = tmp_path / "comprehensive.xlsx"
    wb.save(path)
    return path


class TestComprehensiveXlsx:
    """Tests using a realistic multi-sheet workbook."""

    @pytest.fixture(autouse=True)
    def _convert(self, tmp_path: Path) -> None:
        xlsx_path = _make_comprehensive_xlsx(tmp_path)
        out_dir = tmp_path / "output"
        out_dir.mkdir()
        self.result = import_xlsx(xlsx_path, output_path=out_dir)
        self.content = self.result.markdown_path.read_text()

    def test_converts_successfully(self) -> None:
        """All sheets convert without errors."""
        assert self.result.markdown_path.exists()
        assert len(self.content) > 100

    def test_non_empty_sheets_rendered(self) -> None:
        """All sheets with data get headings."""
        assert "## Summary" in self.content
        assert "## Details" in self.content
        assert "## Trailing" in self.content

    def test_empty_sheet_skipped(self) -> None:
        """Sheet with no data is omitted."""
        assert "## Empty" not in self.content

    def test_integer_values(self) -> None:
        """Integer values rendered correctly."""
        assert "2400000" in self.content

    def test_float_values(self) -> None:
        """Float values rendered with decimals."""
        assert "1800000.5" in self.content

    def test_whole_number_float(self) -> None:
        """Whole-number floats drop the .0."""
        # 150.0 should render as "150", 42.0 as "42"
        assert "| 150" in self.content

    def test_date_formatting(self) -> None:
        """Dates formatted as YYYY-MM-DD."""
        assert "2024-12-31" in self.content
        assert "2025-01-15" in self.content

    def test_datetime_formatting(self) -> None:
        """Datetimes formatted as YYYY-MM-DD HH:MM:SS."""
        assert "2024-12-31 17:00:00" in self.content

    def test_time_formatting(self) -> None:
        """Time values formatted as HH:MM:SS."""
        assert "09:30:00" in self.content
        assert "17:00:00" in self.content

    def test_pipe_chars_escaped(self) -> None:
        """Pipe characters in cell values are escaped."""
        assert "a \\| b" in self.content

    def test_empty_cells(self) -> None:
        """Empty cells produce valid table structure."""
        # Details sheet should have proper pipe table with blanks
        lines = [line for line in self.content.split("\n") if line.startswith("|")]
        # Every data line should have correct pipe count
        for line in lines:
            assert line.endswith("|")

    def test_trailing_empty_rows_stripped(self) -> None:
        """Trailing empty rows in sheets are not rendered."""
        # Trailing sheet should only have header + 2 data rows (not 5)
        trailing_section = self.content.split("## Trailing")[1]
        # Count data rows (lines starting with |, excluding separator)
        data_lines = [
            line
            for line in trailing_section.strip().split("\n")
            if line.startswith("|") and "---" not in line
        ]
        assert len(data_lines) == 3  # header + alpha + beta

    def test_no_images(self) -> None:
        """XLSX import produces no images."""
        assert self.result.images == []
