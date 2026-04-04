"""XLSX to Markdown import converter."""

from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook
from rich.console import Console

from leafpress.exceptions import XlsxImportError
from leafpress.importer.base import (
    ImportResult,
    postprocess_markdown,
    resolve_output_path,
    rows_to_pipe_table,
)

console = Console()


def import_xlsx(
    xlsx_path: Path,
    output_path: Path | None = None,
) -> ImportResult:
    """Convert an XLSX file to Markdown tables.

    Each worksheet becomes a section with a ``## Sheet Name`` heading
    followed by a pipe-style Markdown table.  The first row of each
    sheet is treated as the header row.

    Args:
        xlsx_path: Path to the input .xlsx file.
        output_path: Output .md file path or directory.  If None, uses
                      xlsx_path stem + .md in the same directory.

    Returns:
        ImportResult with the path to the generated Markdown file.

    Raises:
        XlsxImportError: If the input file is invalid or conversion fails.
    """
    if not xlsx_path.exists():
        raise XlsxImportError(f"File not found: {xlsx_path}")
    if xlsx_path.suffix.lower() != ".xlsx":
        raise XlsxImportError(f"Not a .xlsx file: {xlsx_path}")

    md_path = resolve_output_path(xlsx_path, output_path)

    with console.status("[bold blue]Converting XLSX to Markdown..."):
        try:
            wb = load_workbook(xlsx_path, data_only=True)
        except Exception as e:
            raise XlsxImportError(f"Failed to open XLSX: {e}") from e

        warnings: list[str] = []
        sections: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            _collect_sheet_warnings(ws, sheet_name, warnings)
            table_md = _sheet_to_markdown(ws)
            if table_md:
                sections.append(f"## {sheet_name}\n\n{table_md}")

        wb.close()

    if not sections:
        warnings.append("Workbook has no non-empty sheets — output will be blank")

    markdown = "\n\n".join(sections)
    markdown = postprocess_markdown(markdown)

    md_path.parent.mkdir(parents=True, exist_ok=True)
    md_path.write_text(markdown, encoding="utf-8")

    return ImportResult(
        markdown_path=md_path,
        images=[],
        warnings=warnings,
    )


def _collect_sheet_warnings(ws, sheet_name: str, warnings: list[str]) -> None:
    """Check a worksheet for features that can't be fully converted."""
    # Merged cells — only the top-left value is preserved
    merged = list(ws.merged_cells.ranges)
    if merged:
        warnings.append(
            f"Sheet '{sheet_name}' has {len(merged)} merged cell region(s) "
            f"— only the top-left cell value of each is kept"
        )

    # Embedded images
    image_count = len(ws._images)
    if image_count:
        warnings.append(f"Sheet '{sheet_name}' has {image_count} embedded image(s) — not extracted")

    # Embedded charts
    chart_count = len(ws._charts)
    if chart_count:
        warnings.append(f"Sheet '{sheet_name}' has {chart_count} embedded chart(s) — not extracted")


def _sheet_to_markdown(ws) -> str:
    """Convert a worksheet to a pipe-style Markdown table."""
    rows: list[list[str]] = []
    for row in ws.iter_rows():
        cells = [_cell_to_str(cell.value) for cell in row]
        rows.append(cells)

    # Strip fully-empty trailing rows
    while rows and all(c == "" for c in rows[-1]):
        rows.pop()

    return rows_to_pipe_table(rows)


def _cell_to_str(value: object) -> str:
    """Convert a cell value to a display string."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value)
    return str(value)
